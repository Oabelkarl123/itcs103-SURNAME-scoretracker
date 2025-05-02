import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
import os

EXCEL_FILE = "data2.xlsx"
PASS_MARK = 50
selected_item_id = None

if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Scores"
    ws.append(["ID", "Name", "Score", "Status"])
    wb.save(EXCEL_FILE)

def load_data():
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    return [row for row in ws.iter_rows(min_row=2, values_only=True)]

def refresh_listbox():
    listbox.delete(*listbox.get_children())
    for record in load_data():
        listbox.insert("", tk.END, values=record)

def clear_entries():
    global selected_item_id
    id_entry.config(state='normal')
    id_entry.delete(0, tk.END)
    name_entry.delete(0, tk.END)
    score_entry.delete(0, tk.END)
    selected_item_id = None

def add_entry():
    id_val = id_entry.get().strip()
    name_val = name_entry.get().strip()
    score_text = score_entry.get().strip()

    if not (id_val and name_val and score_text):
        messagebox.showwarning("Missing Data", "All fields are required.")
        return

    try:
        score = int(score_text)
    except ValueError:
        messagebox.showerror("Invalid Score", "Score must be a number.")
        return

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == id_val:
            messagebox.showerror("Duplicate ID", "An entry with this ID already exists.")
            return

    status = "Pass" if score >= PASS_MARK else "Fail"
    ws.append([id_val, name_val, score, status])
    wb.save(EXCEL_FILE)
    refresh_listbox()
    clear_entries()

def on_select(event):
    global selected_item_id
    selected = listbox.selection()
    if not selected:
        return
    item = listbox.item(selected[0])
    id_val, name_val, score_val, _ = item['values']
    selected_item_id = id_val

    id_entry.config(state='normal')
    id_entry.delete(0, tk.END)
    id_entry.insert(0, id_val)
    id_entry.config(state='disabled')

    name_entry.delete(0, tk.END)
    name_entry.insert(0, name_val)
    score_entry.delete(0, tk.END)
    score_entry.insert(0, score_val)

def update_entry():
    global selected_item_id
    if not selected_item_id:
        messagebox.showwarning("Select Entry", "Select an entry to update.")
        return

    new_id = selected_item_id
    new_name = name_entry.get().strip()
    new_score_text = score_entry.get().strip()

    if not (new_name and new_score_text):
        messagebox.showwarning("Missing Data", "All fields are required.")
        return

    try:
        new_score = int(new_score_text)
    except ValueError:
        messagebox.showerror("Invalid Score", "Score must be a number.")
        return

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == str(selected_item_id):
            row[1].value = new_name
            row[2].value = new_score
            row[3].value = "Pass" if new_score >= PASS_MARK else "Fail"
            break

    wb.save(EXCEL_FILE)
    refresh_listbox()
    clear_entries()

def delete_entry():
    global selected_item_id
    if not selected_item_id:
        messagebox.showwarning("Select Entry", "Select an entry to delete.")
        return

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    for row in list(ws.iter_rows(min_row=2)):
        if str(row[0].value) == str(selected_item_id):
            ws.delete_rows(row[0].row)
            break

    wb.save(EXCEL_FILE)
    refresh_listbox()
    clear_entries()

# === GUI Setup ===
root = tk.Tk()
root.title("Student Score Tracker")
root.geometry("750x600")
root.configure(bg="lavender")

style = ttk.Style(root)
style.theme_use("clam")
style.configure("TButton", padding=6, relief="flat", background="slateblue", foreground="white",
                font=("Segoe UI", 10, "bold"))
style.map("TButton", background=[("active", "darkslateblue")])
style.configure("TLabel", background="lavender", font=("Segoe UI", 10))
style.configure("Treeview", font=("Segoe UI", 10), rowheight=25)
style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"))

# === Header Bar ===
header = tk.Frame(root, bg="mediumorchid", height=60) 
header.pack(fill="x")
tk.Label(header, text="Student Score Tracker", bg="mediumorchid", fg="white",
         font=("Segoe UI", 16, "bold")).pack(pady=10)

# === Form Frame ===
form_frame = tk.Frame(root, bg="lavender")
form_frame.pack(pady=20)

ttk.Label(form_frame, text="ID:").grid(row=0, column=0, sticky="e", padx=10, pady=8)
ttk.Label(form_frame, text="Name:").grid(row=1, column=0, sticky="e", padx=10, pady=8)
ttk.Label(form_frame, text="Score:").grid(row=2, column=0, sticky="e", padx=10, pady=8)

id_entry = ttk.Entry(form_frame, width=40)
name_entry = ttk.Entry(form_frame, width=40)
score_entry = ttk.Entry(form_frame, width=40)

id_entry.grid(row=0, column=1, padx=10, pady=8)
name_entry.grid(row=1, column=1, padx=10, pady=8)
score_entry.grid(row=2, column=1, padx=10, pady=8)

# === Button Frame ===
btn_frame = tk.Frame(root, bg="lavender")
btn_frame.pack(pady=10)

ttk.Button(btn_frame, text="Add", command=add_entry).grid(row=0, column=0, padx=10)
ttk.Button(btn_frame, text="Update", command=update_entry).grid(row=0, column=1, padx=10)
ttk.Button(btn_frame, text="Delete", command=delete_entry).grid(row=0, column=2, padx=10)
ttk.Button(btn_frame, text="Clear", command=clear_entries).grid(row=0, column=3, padx=10)

# === Treeview Frame ===
tree_frame = tk.Frame(root, bg="lavender")
tree_frame.pack(pady=20, padx=20, fill='both', expand=True)

columns = ("ID", "Name", "Score", "Status")
listbox = ttk.Treeview(tree_frame, columns=columns, show="headings")
for col in columns:
    listbox.heading(col, text=col)
    listbox.column(col, width=150, anchor="center")

listbox.pack(fill="both", expand=True)
listbox.bind("<<TreeviewSelect>>", on_select)

refresh_listbox()
root.mainloop()
